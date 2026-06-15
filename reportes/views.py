from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from alquileres.models import Alquiler
from inventario.models import Herramienta, Categoria
from clientes.models import Cliente
from multas.models import Multa
from django.db.models import Sum, Count
import io
from datetime import datetime


@login_required
def lista(request):
    context = {
        'total_alquileres': Alquiler.objects.count(),
        'total_herramientas': Herramienta.objects.filter(activo=True).count(),
        'total_clientes': Cliente.objects.count(),
        'total_multas': Multa.objects.filter(estado='pendiente').aggregate(
            total=Sum('monto_total'))['total'] or 0,
    }
    return render(request, 'reportes/lista.html', context)


def aplicar_filtros_alquileres(request):
    # Aplica filtros a los alquileres según los parámetros del request
    alquileres = Alquiler.objects.select_related('cliente__usuario').order_by('-fecha_creacion')

    estado = request.GET.get('estado', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    canal = request.GET.get('canal', '')

    if estado:
        alquileres = alquileres.filter(estado=estado)
    if fecha_inicio:
        alquileres = alquileres.filter(fecha_inicio__gte=fecha_inicio)
    if fecha_fin:
        alquileres = alquileres.filter(fecha_fin__lte=fecha_fin)
    if canal:
        alquileres = alquileres.filter(canal=canal)

    return alquileres


def aplicar_filtros_inventario(request):
    # Aplica filtros al inventario según los parámetros del request
    herramientas = Herramienta.objects.filter(activo=True).select_related('categoria')

    estado = request.GET.get('estado', '')
    categoria_id = request.GET.get('categoria', '')

    if estado:
        herramientas = herramientas.filter(estado=estado)
    if categoria_id:
        herramientas = herramientas.filter(categoria_id=categoria_id)

    return herramientas


def aplicar_filtros_multas(request):
    # Aplica filtros a las multas según los parámetros del request
    multas = Multa.objects.select_related('cliente__usuario', 'alquiler').order_by('-fecha_creacion')

    estado = request.GET.get('estado', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')

    if estado:
        multas = multas.filter(estado=estado)
    if fecha_inicio:
        multas = multas.filter(fecha_vencimiento__gte=fecha_inicio)
    if fecha_fin:
        multas = multas.filter(fecha_vencimiento__lte=fecha_fin)

    return multas


@login_required
def reporte_inventario(request):
    # Vista de reporte de inventario con filtros
    herramientas = aplicar_filtros_inventario(request)
    categorias = Categoria.objects.filter(activo=True)

    context = {
        'herramientas': herramientas,
        'categorias': categorias,
        'total': herramientas.count(),
        'estado': request.GET.get('estado', ''),
        'categoria': request.GET.get('categoria', ''),
    }
    return render(request, 'reportes/inventario.html', context)


@login_required
def reporte_alquileres(request):
    # Vista de reporte de alquileres con filtros
    alquileres = aplicar_filtros_alquileres(request)

    context = {
        'alquileres': alquileres,
        'total': alquileres.count(),
        'total_ingresos': alquileres.aggregate(total=Sum('total'))['total'] or 0,
        'estado': request.GET.get('estado', ''),
        'fecha_inicio': request.GET.get('fecha_inicio', ''),
        'fecha_fin': request.GET.get('fecha_fin', ''),
        'canal': request.GET.get('canal', ''),
    }
    return render(request, 'reportes/alquileres.html', context)


@login_required
def reporte_multas(request):
    # Vista de reporte de multas con filtros
    multas = aplicar_filtros_multas(request)

    context = {
        'multas': multas,
        'total': multas.count(),
        'total_monto': multas.aggregate(total=Sum('monto_total'))['total'] or 0,
        'estado': request.GET.get('estado', ''),
        'fecha_inicio': request.GET.get('fecha_inicio', ''),
        'fecha_fin': request.GET.get('fecha_fin', ''),
    }
    return render(request, 'reportes/multas.html', context)


@login_required
def reporte_inventario_pdf(request):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    herramientas = aplicar_filtros_inventario(request)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph('Reporte de Inventario - SCAE', styles['Title']))
    elements.append(Paragraph(f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
    elements.append(Paragraph(f'Total herramientas: {herramientas.count()}', styles['Normal']))
    elements.append(Spacer(1, 20))

    data = [['Código', 'Nombre', 'Categoría', 'Precio/Día', 'Total', 'Disponible', 'Alquilada', 'Estado']]
    for h in herramientas:
        data.append([
            h.codigo, h.nombre, h.categoria.nombre,
            f'${h.precio_diario:,.0f}',
            str(h.cantidad_total),
            str(h.cantidad_disponible),
            str(h.cantidad_alquilada),
            h.get_estado_display(),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4361ee')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="inventario.pdf"'
    return response


@login_required
def reporte_inventario_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    herramientas = aplicar_filtros_inventario(request)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'

    headers = ['Código', 'Nombre', 'Categoría', 'Precio/Día', 'Total', 'Disponible', 'Alquilada', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4361ee', end_color='4361ee', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for h in herramientas:
        ws.append([
            h.codigo, h.nombre, h.categoria.nombre,
            float(h.precio_diario),
            h.cantidad_total, h.cantidad_disponible,
            h.cantidad_alquilada, h.get_estado_display(),
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 5

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventario.xlsx"'
    return response


@login_required
def reporte_alquileres_pdf(request):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    alquileres = aplicar_filtros_alquileres(request)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph('Reporte de Alquileres - SCAE', styles['Title']))
    elements.append(Paragraph(f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
    elements.append(Paragraph(f'Total alquileres: {alquileres.count()}', styles['Normal']))
    elements.append(Spacer(1, 20))

    data = [['N° Alquiler', 'Cliente', 'Fecha Inicio', 'Fecha Fin', 'Total', 'Estado', 'Canal']]
    for a in alquileres:
        data.append([
            a.numero, a.cliente.nombre_completo,
            str(a.fecha_inicio), str(a.fecha_fin),
            f'${a.total:,.0f}',
            a.get_estado_display(),
            a.get_canal_display(),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4361ee')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="alquileres.pdf"'
    return response


@login_required
def reporte_alquileres_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    alquileres = aplicar_filtros_alquileres(request)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Alquileres'

    headers = ['N° Alquiler', 'Cliente', 'Fecha Inicio', 'Fecha Fin', 'Total', 'Estado', 'Canal']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4361ee', end_color='4361ee', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for a in alquileres:
        ws.append([
            a.numero, a.cliente.nombre_completo,
            str(a.fecha_inicio), str(a.fecha_fin),
            float(a.total),
            a.get_estado_display(),
            a.get_canal_display(),
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 5

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="alquileres.xlsx"'
    return response


@login_required
def reporte_multas_pdf(request):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    multas = aplicar_filtros_multas(request)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph('Reporte de Multas - SCAE', styles['Title']))
    elements.append(Paragraph(f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
    elements.append(Paragraph(f'Total multas: {multas.count()}', styles['Normal']))
    elements.append(Spacer(1, 20))

    data = [['Alquiler', 'Cliente', 'Fecha Vencimiento', 'Días Retraso', 'Monto', 'Estado']]
    for m in multas:
        data.append([
            m.alquiler.numero, m.cliente.nombre_completo,
            str(m.fecha_vencimiento),
            str(m.dias_retraso),
            f'${m.monto_total:,.0f}',
            m.get_estado_display(),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e63946')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="multas.pdf"'
    return response


@login_required
def reporte_multas_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    multas = aplicar_filtros_multas(request)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Multas'

    headers = ['Alquiler', 'Cliente', 'Fecha Vencimiento', 'Días Retraso', 'Monto', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='e63946', end_color='e63946', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for m in multas:
        ws.append([
            m.alquiler.numero, m.cliente.nombre_completo,
            str(m.fecha_vencimiento),
            m.dias_retraso,
            float(m.monto_total),
            m.get_estado_display(),
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 5

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="multas.xlsx"'
    return response